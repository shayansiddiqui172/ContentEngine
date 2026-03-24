"""Write enriched creator and post data to living Excel spreadsheets.

DB1_Creator_Profiles.xlsx — one row per creator
DB2_Post_Analysis.xlsx    — one row per post

On each run:
  DB1: match on LINKEDIN LINK. Update scraped fields; preserve manual fields.
  DB2: match on LINK TO POST. Skip existing posts; only append new ones.
"""

import logging
import os
from datetime import datetime

from scripts.scraper.config import (
    DB1_XLSX_PATH,
    DB2_XLSX_PATH,
    DB1_COLUMNS,
    DB2_COLUMNS,
    DB1_DESCRIPTIONS,
    DB2_DESCRIPTIONS,
    follower_count_range,
)

logger = logging.getLogger(__name__)

# ── Styling constants ──────────────────────────────────────────────────────────
HEADER_BG = "2A2927"
HEADER_FG = "FFFFFF"
DESC_BG = "FFFACD"   # light yellow

GREEN_BG = "C6EFCE"
GREEN_FG = "276221"
AMBER_BG = "FFEB9C"
AMBER_FG = "9C5700"
RED_BG   = "FFC7CE"
RED_FG   = "9C0006"

# ── Field protection rules ─────────────────────────────────────────────────────
# DB1: these columns are manually maintained — never overwrite on update
DB1_PROTECTED = {
    "OVERALL NOTES",
    "COLLABORATION POTENTIAL?",
    "CONNECTION STATUS",
    "HAVE INTERACTED W/CONTENT?",
    "HAVE DMED YET?",
    "NOTES FOR RELATIONSHIP HISTORY",
}

# DB2: skip entire row if post already exists (preserves manual edits)
# (no partial update for posts — entire row is immutable once written)

# ── Score columns to colour-code (1-5 integer scale) ──────────────────────────
SCORE_COLS_DB1 = {"CREDIBILITY (1-5)"}
SCORE_COLS_DB2 = {"HOOK STRENGTH (1-5)", "SHAREABILITY (1-5)", "ENGAGEMENT SCORE (1-5)"}


# ── Utility helpers ────────────────────────────────────────────────────────────

def _yn(value) -> str:
    """Normalise a bool/string to 'Y' or 'N'."""
    if value is True or (isinstance(value, str) and value.strip().upper() in ("Y", "YES", "TRUE")):
        return "Y"
    if value is False or (isinstance(value, str) and value.strip().upper() in ("N", "NO", "FALSE")):
        return "N"
    return str(value) if value else ""


def _fmt_date(value) -> str:
    """Convert ISO datetime string to YYYY-MM-DD for the spreadsheet."""
    if not value:
        return ""
    s = str(value)
    if "T" in s:
        try:
            return datetime.fromisoformat(s.replace("Z", "+00:00")).strftime("%Y-%m-%d")
        except (ValueError, TypeError):
            pass
    return s


def _compute_posting_frequency(posts: list[dict]) -> str | None:
    """Derive posting frequency bucket from post publish dates."""
    if not posts:
        return None
    dates = []
    for p in posts:
        pub = p.get("publishedAt")
        if pub:
            try:
                dt = datetime.fromisoformat(str(pub).replace("Z", "+00:00"))
                dates.append(dt)
            except (ValueError, TypeError):
                pass
    if len(dates) < 2:
        return None
    dates.sort()
    span_days = (dates[-1] - dates[0]).days
    if span_days == 0:
        return None
    rate = len(dates) / span_days  # posts per day
    if rate >= 0.85:
        return "Daily"
    if rate >= 0.4:
        return "3-4/week"
    if rate >= 0.14:
        return "1-2/week"
    return "< a month"


# ── Row builders ───────────────────────────────────────────────────────────────

def _creator_to_db1_row(creator: dict) -> dict:
    """Map internal creator dict → DB1 column dict."""
    follower_count = creator.get("followerCount")
    posts = creator.get("posts", [])

    fc_range = creator.get("followerCountRange") or follower_count_range(follower_count)

    growth_stage = creator.get("growthStage")
    if not growth_stage and follower_count is not None:
        growth_stage = "Established Creator" if follower_count >= 50_000 else "Emerging Creator"

    posting_freq = creator.get("postingFrequency") or _compute_posting_frequency(posts)

    # crossPlatformCount: use stored value or fall back to counting has* flags from old schema
    cross_platform = creator.get("crossPlatformCount")
    if cross_platform is None:
        cross_platform = sum(
            1 for f in ("hasTwitter", "hasSubstack", "hasYoutube", "hasPodcast")
            if creator.get(f)
        )

    # Collaboration potential: may be None, bool, or string
    collab = creator.get("collaborationPotential")
    if collab is True:
        collab = "Y"
    elif collab is False:
        collab = "N"
    elif collab is None:
        collab = ""

    return {
        "NAME":                           creator.get("name") or creator.get("fullName") or "",
        "BIO":                            creator.get("bio") or "",
        "FOLLOWER COUNT (RANGE)":         fc_range or "",
        "FOLLOWER COUNT (#)":             follower_count if follower_count is not None else "",
        "LOCATION":                       creator.get("location") or "",
        "FIRM AFFILIATION":               creator.get("firmAffiliation") or creator.get("firmOrCompany") or "",
        "DATE ADDED":                     _fmt_date(creator.get("dateAdded") or creator.get("followerCountUpdatedAt")),
        "OVERALL NOTES":                  creator.get("overallNotes") or "",
        "PRIMARY ROLE":                   creator.get("primaryRole") or "",
        "CONTENT NICHE":                  creator.get("contentNiche") or "",
        "GROWTH STAGE":                   growth_stage or "",
        "GEOGRAPHY FOCUS":                creator.get("geographyFocus") or "",
        "COUNT OF CROSS-PLATFORM PRESENCE": cross_platform if cross_platform is not None else 0,
        "TOP VOICE STYLE":                creator.get("topVoiceStyle") or creator.get("voiceStyle") or "",
        "CREDIBILITY (1-5)":              creator.get("credibility") or creator.get("credibilityScore") or "",
        "COLLABORATION POTENTIAL?":       collab,
        "POSTING FREQUENCY":              posting_freq or "",
        "CONNECTION STATUS":              creator.get("connectionStatus") or "",
        "HAVE INTERACTED W/CONTENT?":     _yn(creator.get("haveInteracted") or creator.get("hasInteractedWithContent")),
        "HAVE DMED YET?":                 _yn(creator.get("haveDMed") or creator.get("hasDMedOrMet")),
        "NOTES FOR RELATIONSHIP HISTORY": creator.get("relationshipNotes") or "",
        "LINKEDIN LINK":                  creator.get("linkedinUrl") or "",
    }


def _post_to_db2_row(post: dict, creator_name: str) -> dict:
    """Map internal post dict → DB2 column dict."""
    return {
        "NAME":                          creator_name,
        "POST DATE":                     _fmt_date(post.get("publishedAt")),
        "POST FORMAT":                   post.get("postFormat") or post.get("format") or "",
        "PRIMARY TOPIC":                 post.get("primaryTopic") or "",
        "TOPIC/SUBJECT OF THE POST":     post.get("topicSubject") or post.get("keyInsight") or post.get("summary") or "",
        "DOES IT CONTAIN DATA/STATS?":   _yn(post.get("containsData")),
        "RELEVANCE (to startups/VC)":    post.get("relevance") or "",
        "HOOK":                          post.get("hook") or "",
        "HOOK STYLE":                    post.get("hookStyle") or "",
        "HOOK STRENGTH (1-5)":           post.get("hookStrength") or "",
        "TONE":                          post.get("tone") or "",
        "DOES IT CONTAIN CTA?":          _yn(post.get("containsCTA")),
        "WHAT CTA?":                     post.get("ctaText") or post.get("ctaType") or "",
        "LIKES":                         post.get("reactions") or 0,
        "COMMENTS":                      post.get("comments") or 0,
        "SHARES/REPOSTS":                post.get("reposts") or 0,
        "SHAREABILITY (1-5)":            post.get("shareability") or "",
        "ENGAGEMENT SCORE (1-5)":        post.get("engagementScore") or "",
        "POST PERFORMANCE":              post.get("postPerformance") or "",
        "TOPIC/SUBJECT POPULARITY":      post.get("topicPopularity") or "",
        "POST FREQUENCY":                post.get("postFrequencyAssessment") or "",
        "LINK TO POST":                  post.get("postUrl") or "",
        "Does this post work or not work?": post.get("postWorkAnalysis") or post.get("whatItDoesWell") or "",
    }


# ── Formatting helpers ─────────────────────────────────────────────────────────

def _style_header_row(ws, columns: list[str]):
    from openpyxl.styles import PatternFill, Font, Alignment
    fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    font = Font(bold=True, color=HEADER_FG, size=11)
    for col_idx, name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = name
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _style_description_row(ws, columns: list[str], descriptions: dict):
    from openpyxl.styles import PatternFill, Font, Alignment
    fill = PatternFill(start_color=DESC_BG, end_color=DESC_BG, fill_type="solid")
    font = Font(italic=True, size=9, color="666666")
    for col_idx, name in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = descriptions.get(name, "")
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _apply_score_coloring(ws, columns: list[str], score_cols: set, data_start: int, data_end: int):
    from openpyxl.styles import PatternFill, Font
    green_fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
    amber_fill = PatternFill(start_color=AMBER_BG, end_color=AMBER_BG, fill_type="solid")
    red_fill   = PatternFill(start_color=RED_BG,   end_color=RED_BG,   fill_type="solid")
    green_font = Font(color=GREEN_FG)
    amber_font = Font(color=AMBER_FG)
    red_font   = Font(color=RED_FG)

    for col_idx, name in enumerate(columns, 1):
        if name not in score_cols:
            continue
        for row in range(data_start, data_end + 1):
            cell = ws.cell(row=row, column=col_idx)
            val = cell.value
            if not isinstance(val, (int, float)):
                continue
            if val >= 4:
                cell.fill = green_fill
                cell.font = green_font
            elif val == 3:
                cell.fill = amber_fill
                cell.font = amber_font
            else:
                cell.fill = red_fill
                cell.font = red_font


def _apply_performance_coloring(ws, columns: list[str], data_start: int, data_end: int):
    from openpyxl.styles import PatternFill, Font
    if "POST PERFORMANCE" not in columns:
        return
    col_idx = columns.index("POST PERFORMANCE") + 1
    green_fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
    amber_fill = PatternFill(start_color=AMBER_BG, end_color=AMBER_BG, fill_type="solid")
    red_fill   = PatternFill(start_color=RED_BG,   end_color=RED_BG,   fill_type="solid")
    green_font = Font(color=GREEN_FG)
    amber_font = Font(color=AMBER_FG)
    red_font   = Font(color=RED_FG)

    for row in range(data_start, data_end + 1):
        cell = ws.cell(row=row, column=col_idx)
        val = str(cell.value or "")
        if "Above" in val:
            cell.fill = green_fill
            cell.font = green_font
        elif "Below" in val:
            cell.fill = red_fill
            cell.font = red_font
        elif "Average" in val:
            cell.fill = amber_fill
            cell.font = amber_font


def _auto_size_columns(ws, columns: list[str]):
    from openpyxl.utils import get_column_letter
    for col_idx, name in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(name)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    cell_len = min(len(str(cell.value)), 200)
                    if cell_len > max_len:
                        max_len = cell_len
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)


def _finalize_sheet(ws, columns: list[str], score_cols: set, data_start: int, data_end: int):
    from openpyxl.utils import get_column_letter
    # Freeze header + description rows
    ws.freeze_panes = "A3"
    # Auto-filter on header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    # Row heights
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 42
    # Colour coding
    if data_end >= data_start:
        _apply_score_coloring(ws, columns, score_cols, data_start, data_end)
        _apply_performance_coloring(ws, columns, data_start, data_end)
    # Column widths
    _auto_size_columns(ws, columns)


# ── Main entry point ───────────────────────────────────────────────────────────

def write_spreadsheets(creators: list[dict]):
    """Build DB1 and DB2 rows from creators list, then write/merge Excel files."""
    try:
        import openpyxl  # noqa: F401 — confirm available before building rows
    except ImportError:
        logger.error(
            "openpyxl is not installed. Excel output skipped. "
            "Fix with: pip install openpyxl"
        )
        return

    db1_rows = [_creator_to_db1_row(c) for c in creators]

    db2_rows = []
    for creator in creators:
        name = creator.get("name") or creator.get("fullName") or ""
        for post in creator.get("posts", []):
            db2_rows.append(_post_to_db2_row(post, name))

    _write_db1(db1_rows)
    _write_db2(db2_rows)
    logger.info(f"Spreadsheets written to {os.path.dirname(DB1_XLSX_PATH)}/")


# ── DB1 writer ─────────────────────────────────────────────────────────────────

def _write_db1(new_rows: list[dict]):
    from openpyxl import Workbook, load_workbook

    if os.path.exists(DB1_XLSX_PATH):
        wb = load_workbook(DB1_XLSX_PATH)
        ws = wb.active

        linkedin_col = DB1_COLUMNS.index("LINKEDIN LINK") + 1
        existing_index: dict[str, int] = {}
        for row_num in range(3, ws.max_row + 1):
            cell_val = ws.cell(row=row_num, column=linkedin_col).value
            if cell_val:
                existing_index[str(cell_val).strip()] = row_num

        updated = added = 0
        for row_data in new_rows:
            url = str(row_data.get("LINKEDIN LINK", "")).strip()
            if url and url in existing_index:
                # Overwrite scraped fields only; skip manually-maintained columns
                row_num = existing_index[url]
                for col_idx, col_name in enumerate(DB1_COLUMNS, 1):
                    if col_name not in DB1_PROTECTED:
                        ws.cell(row=row_num, column=col_idx).value = row_data.get(col_name, "")
                updated += 1
            else:
                ws.append([row_data.get(col, "") for col in DB1_COLUMNS])
                added += 1

        data_end = ws.max_row
        _finalize_sheet(ws, DB1_COLUMNS, SCORE_COLS_DB1, 3, data_end)
        wb.save(DB1_XLSX_PATH)
        logger.info(f"DB1: updated {updated} creators, added {added} new → {DB1_XLSX_PATH}")

    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Creator Profiles"
        _style_header_row(ws, DB1_COLUMNS)
        _style_description_row(ws, DB1_COLUMNS, DB1_DESCRIPTIONS)
        for row_data in new_rows:
            ws.append([row_data.get(col, "") for col in DB1_COLUMNS])
        data_end = ws.max_row
        _finalize_sheet(ws, DB1_COLUMNS, SCORE_COLS_DB1, 3, data_end)
        wb.save(DB1_XLSX_PATH)
        logger.info(f"DB1: created fresh with {len(new_rows)} creators → {DB1_XLSX_PATH}")


# ── DB2 writer ─────────────────────────────────────────────────────────────────

def _write_db2(new_rows: list[dict]):
    from openpyxl import Workbook, load_workbook

    if os.path.exists(DB2_XLSX_PATH):
        wb = load_workbook(DB2_XLSX_PATH)
        ws = wb.active

        link_col = DB2_COLUMNS.index("LINK TO POST") + 1
        existing_urls: set[str] = set()
        for row_num in range(3, ws.max_row + 1):
            cell_val = ws.cell(row=row_num, column=link_col).value
            if cell_val:
                existing_urls.add(str(cell_val).strip())

        added = skipped = 0
        for row_data in new_rows:
            url = str(row_data.get("LINK TO POST", "")).strip()
            if url and url in existing_urls:
                skipped += 1
                continue
            ws.append([row_data.get(col, "") for col in DB2_COLUMNS])
            added += 1

        data_end = ws.max_row
        _finalize_sheet(ws, DB2_COLUMNS, SCORE_COLS_DB2, 3, data_end)
        wb.save(DB2_XLSX_PATH)
        logger.info(f"DB2: skipped {skipped} existing posts, added {added} new → {DB2_XLSX_PATH}")

    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Post Analysis"
        _style_header_row(ws, DB2_COLUMNS)
        _style_description_row(ws, DB2_COLUMNS, DB2_DESCRIPTIONS)
        for row_data in new_rows:
            ws.append([row_data.get(col, "") for col in DB2_COLUMNS])
        data_end = ws.max_row
        _finalize_sheet(ws, DB2_COLUMNS, SCORE_COLS_DB2, 3, data_end)
        wb.save(DB2_XLSX_PATH)
        logger.info(f"DB2: created fresh with {len(new_rows)} posts → {DB2_XLSX_PATH}")
